"""
excel_style.py — Apply institutional financial-modeling styling to an openpyxl Workbook.

Implements the canonical sell-side / buy-side modeling conventions:
- Blue inputs, black formulas, green internal links, purple external links
- Calibri 11pt, no gridlines, parentheses for negatives
- Navy section headers, grey subheaders, thin/double-rule subtotals/totals
- Named styles applied via NamedStyle objects (one-time cost, consistent output)

Usage (programmatic):

    from openpyxl import Workbook
    from excel_style import apply_institutional_styles, write_header, write_input, write_formula, write_section, write_subtotal, write_total, set_sheet_defaults

    wb = Workbook()
    apply_institutional_styles(wb)
    ws = wb.active
    set_sheet_defaults(ws)

    write_header(ws, 1, "Pro Forma — Marina Apartments — Q1 2026")
    write_section(ws, 3, "Operating Performance")
    ws.cell(row=5, column=2, value="Revenue").style = "label"
    write_input(ws, "C5", 4500000, fmt="dollar")
    write_formula(ws, "C6", "=C5*0.045", fmt="dollar")
    write_subtotal(ws, "C7", "=SUM(C5:C6)", fmt="dollar")

    wb.save("model.xlsx")

CLI smoke-test:
    python excel_style.py --demo output.xlsx
"""

from __future__ import annotations

import argparse

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Color,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

NAVY = "1F3864"
DEEP_NAVY = "002060"
INK = "222222"
CHARCOAL = "333333"
MID_GREY = "7F7F7F"
SOFT_GREY = "D9D9D9"
PAPER_GREY = "F2F2F2"
INPUT_BLUE = "0000FF"
LINK_GREEN = "00B050"
LINK_PURPLE = "800080"
FLAG_RED = "FF0000"
HIGHLIGHT_YELLOW = "FFF2CC"
FIX_YELLOW = "FFFF00"
WHITE = "FFFFFF"

ACCENT_MCKINSEY_BLUE = "005EB8"
ACCENT_BROOKFIELD_ORANGE = "FF8200"
ACCENT_BURGUNDY = "7B1F2B"


# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

FORMATS = {
    "dollar":       '#,##0;(#,##0);"-"',
    "dollar_k":     '#,##0,;(#,##0,);"-"',      # scaled to thousands
    "dollar_m":     '#,##0.0,,;(#,##0.0,,);"-"',  # scaled to millions
    "currency":     '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)',
    "pct1":         '0.0%;(0.0%);"-"',
    "pct2":         '0.00%;(0.00%);"-"',
    "bps":          '0" bps";(0" bps");"-"',
    "multiple":     '0.00"x";(0.00"x");"-"',
    "years":        '0.0" yrs"',
    "date":         "mmm-yy",
    "date_full":    "mmm-yyyy",
    "sf":           '#,##0" SF"',
    "per_sf":       '$#,##0.00"/SF"',
    "per_unit":     '$#,##0"/unit"',
    "general":      "General",
}


def _font(color: str = INK, bold: bool = False, italic: bool = False, size: int = 11) -> Font:
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def _fill(color: str | None) -> PatternFill | None:
    if color is None:
        return None
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def _side(style: str | None = "thin", color: str = INK) -> Side:
    return Side(border_style=style, color=color)


# ---------------------------------------------------------------------------
# Named styles
# ---------------------------------------------------------------------------

def _build_named_styles() -> list[NamedStyle]:
    """Construct the institutional NamedStyle set. Called once per workbook."""
    styles: list[NamedStyle] = []

    # --- Inputs (blue font, yellow fill) ---
    for suffix, fmt in [
        ("dollar",   FORMATS["dollar"]),
        ("dollar_m", FORMATS["dollar_m"]),
        ("pct1",     FORMATS["pct1"]),
        ("pct2",     FORMATS["pct2"]),
        ("multiple", FORMATS["multiple"]),
        ("bps",      FORMATS["bps"]),
        ("years",    FORMATS["years"]),
        ("date",     FORMATS["date"]),
        ("general",  FORMATS["general"]),
        ("sf",       FORMATS["sf"]),
        ("per_sf",   FORMATS["per_sf"]),
        ("per_unit", FORMATS["per_unit"]),
    ]:
        s = NamedStyle(name=f"input_{suffix}")
        s.font = _font(INPUT_BLUE)
        s.fill = _fill(HIGHLIGHT_YELLOW)
        s.alignment = Alignment(horizontal="right", vertical="center")
        s.number_format = fmt
        styles.append(s)

    # --- Formulas (black font, no fill) ---
    for suffix, fmt in [
        ("dollar",   FORMATS["dollar"]),
        ("dollar_m", FORMATS["dollar_m"]),
        ("pct1",     FORMATS["pct1"]),
        ("pct2",     FORMATS["pct2"]),
        ("multiple", FORMATS["multiple"]),
        ("bps",      FORMATS["bps"]),
        ("years",    FORMATS["years"]),
        ("general",  FORMATS["general"]),
        ("sf",       FORMATS["sf"]),
        ("per_sf",   FORMATS["per_sf"]),
        ("per_unit", FORMATS["per_unit"]),
    ]:
        s = NamedStyle(name=f"formula_{suffix}")
        s.font = _font(INK)
        s.alignment = Alignment(horizontal="right", vertical="center")
        s.number_format = fmt
        styles.append(s)

    # --- Cross-sheet link (green font) ---
    for suffix, fmt in [("dollar", FORMATS["dollar"]), ("pct1", FORMATS["pct1"])]:
        s = NamedStyle(name=f"link_internal_{suffix}")
        s.font = _font(LINK_GREEN)
        s.alignment = Alignment(horizontal="right", vertical="center")
        s.number_format = fmt
        styles.append(s)

    # --- External link (purple font) ---
    s = NamedStyle(name="link_external_dollar")
    s.font = _font(LINK_PURPLE)
    s.alignment = Alignment(horizontal="right", vertical="center")
    s.number_format = FORMATS["dollar"]
    styles.append(s)

    # --- Flag (bold red on yellow) ---
    s = NamedStyle(name="flag_cell")
    s.font = _font(FLAG_RED, bold=True)
    s.fill = _fill(FIX_YELLOW)
    s.alignment = Alignment(horizontal="right", vertical="center")
    s.number_format = FORMATS["dollar"]
    styles.append(s)

    # --- Section header (white on navy) ---
    s = NamedStyle(name="section_header")
    s.font = _font(WHITE, bold=True)
    s.fill = _fill(NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    # --- Page header banner (white on navy, large) ---
    s = NamedStyle(name="page_header")
    s.font = _font(WHITE, bold=True, size=12)
    s.fill = _fill(NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    # --- Subheader (bold on grey) ---
    s = NamedStyle(name="subheader")
    s.font = _font(INK, bold=True)
    s.fill = _fill(SOFT_GREY)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    # --- Subtotal (bold + thin top border) ---
    for suffix, fmt in [("dollar", FORMATS["dollar"]), ("pct1", FORMATS["pct1"])]:
        s = NamedStyle(name=f"subtotal_{suffix}")
        s.font = _font(INK, bold=True)
        s.alignment = Alignment(horizontal="right", vertical="center")
        s.number_format = fmt
        s.border = Border(top=_side("thin"))
        styles.append(s)

    # --- Total (bold + thin top + double bottom) ---
    for suffix, fmt in [("dollar", FORMATS["dollar"]), ("pct1", FORMATS["pct1"]), ("multiple", FORMATS["multiple"])]:
        s = NamedStyle(name=f"total_{suffix}")
        s.font = _font(INK, bold=True)
        s.alignment = Alignment(horizontal="right", vertical="center")
        s.number_format = fmt
        s.border = Border(top=_side("thin"), bottom=_side("double"))
        styles.append(s)

    # --- Label (left-aligned text) ---
    s = NamedStyle(name="label")
    s.font = _font(INK)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    s = NamedStyle(name="label_bold")
    s.font = _font(INK, bold=True)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    # --- Units cell (italic grey, right aligned) ---
    s = NamedStyle(name="units_cell")
    s.font = _font(MID_GREY, italic=True)
    s.alignment = Alignment(horizontal="right", vertical="center")
    styles.append(s)

    # --- Date column header ---
    s = NamedStyle(name="date_header")
    s.font = _font(INK, bold=True)
    s.alignment = Alignment(horizontal="center", vertical="center")
    s.number_format = FORMATS["date"]
    styles.append(s)

    # --- Note (footnote, 9pt italic grey) ---
    s = NamedStyle(name="note_cell")
    s.font = _font(MID_GREY, italic=True, size=9)
    s.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(s)

    return styles


def apply_institutional_styles(wb: Workbook) -> None:
    """Register all institutional NamedStyles on the workbook. Call once after Workbook()."""
    # openpyxl >= 3.1 returns a list of style names (strings) from wb.named_styles.
    existing = set(wb.named_styles) if hasattr(wb, "named_styles") else set()
    for style in _build_named_styles():
        if style.name not in existing:
            wb.add_named_style(style)


# ---------------------------------------------------------------------------
# Sheet-level setup
# ---------------------------------------------------------------------------

def set_sheet_defaults(ws: Worksheet, title: str | None = None) -> None:
    """Apply the institutional sheet-level defaults: no gridlines, landscape, fit-to-1-wide."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7
    ws.oddHeader.left.text = title or ws.title
    ws.oddFooter.left.text = "&F"
    ws.oddFooter.center.text = "&A"
    ws.oddFooter.right.text = "Page &P of &N"

    # FAST column structure defaults
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 13


# ---------------------------------------------------------------------------
# Writer helpers
# ---------------------------------------------------------------------------

def _coerce_cell(ws: Worksheet, ref_or_rc):
    if isinstance(ref_or_rc, tuple):
        row, col = ref_or_rc
        return ws.cell(row=row, column=col)
    return ws[ref_or_rc]


def write_header(ws: Worksheet, row: int, title: str, span_cols: int = 12) -> None:
    """Write a navy/white page-title banner across row {row}."""
    cell = ws.cell(row=row, column=2, value=title)
    cell.style = "page_header"
    ws.row_dimensions[row].height = 26
    # Apply fill across the banner range for print appearance
    for col in range(3, 2 + span_cols):
        c = ws.cell(row=row, column=col)
        c.style = "page_header"


def write_section(ws: Worksheet, row: int, title: str, span_cols: int = 12) -> None:
    """Write a navy section header at {row}."""
    cell = ws.cell(row=row, column=2, value=title)
    cell.style = "section_header"
    ws.row_dimensions[row].height = 21
    for col in range(3, 2 + span_cols):
        c = ws.cell(row=row, column=col)
        c.style = "section_header"


def write_subheader(ws: Worksheet, row: int, title: str, span_cols: int = 12) -> None:
    cell = ws.cell(row=row, column=2, value=title)
    cell.style = "subheader"
    for col in range(3, 2 + span_cols):
        c = ws.cell(row=row, column=col)
        c.style = "subheader"


def write_label(ws: Worksheet, ref, text: str, bold: bool = False) -> None:
    c = _coerce_cell(ws, ref)
    c.value = text
    c.style = "label_bold" if bold else "label"


def write_units(ws: Worksheet, ref, units: str) -> None:
    c = _coerce_cell(ws, ref)
    c.value = units
    c.style = "units_cell"


def write_input(ws: Worksheet, ref, value, fmt: str = "dollar") -> None:
    """Write a hard-coded input. Blue font, yellow fill."""
    c = _coerce_cell(ws, ref)
    c.value = value
    style = f"input_{fmt}"
    c.style = style


def write_formula(ws: Worksheet, ref, formula: str, fmt: str = "dollar") -> None:
    """Write a formula (must start with =). Black font."""
    c = _coerce_cell(ws, ref)
    c.value = formula if formula.startswith("=") else f"={formula}"
    style = f"formula_{fmt}"
    c.style = style


def write_link_internal(ws: Worksheet, ref, formula: str, fmt: str = "dollar") -> None:
    """Write a cross-sheet link. Green font."""
    c = _coerce_cell(ws, ref)
    c.value = formula if formula.startswith("=") else f"={formula}"
    c.style = f"link_internal_{fmt}"


def write_subtotal(ws: Worksheet, ref, value_or_formula, fmt: str = "dollar") -> None:
    c = _coerce_cell(ws, ref)
    if isinstance(value_or_formula, str):
        c.value = value_or_formula if value_or_formula.startswith("=") else f"={value_or_formula}"
    else:
        c.value = value_or_formula
    c.style = f"subtotal_{fmt}"


def write_total(ws: Worksheet, ref, value_or_formula, fmt: str = "dollar") -> None:
    c = _coerce_cell(ws, ref)
    if isinstance(value_or_formula, str):
        c.value = value_or_formula if value_or_formula.startswith("=") else f"={value_or_formula}"
    else:
        c.value = value_or_formula
    c.style = f"total_{fmt}"


def write_note(ws: Worksheet, row: int, text: str, col: int = 2) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.style = "note_cell"


# ---------------------------------------------------------------------------
# Demo / smoke test
# ---------------------------------------------------------------------------

def _build_demo(path: str) -> None:
    wb = Workbook()
    apply_institutional_styles(wb)
    ws = wb.active
    ws.title = "Pro Forma"
    set_sheet_defaults(ws, "Pro Forma — Marina Apartments — Q1 2026")

    write_header(ws, 1, "Pro Forma — Marina Apartments — Q1 2026", span_cols=8)

    write_section(ws, 3, "Operating Performance — T-12 vs. UW", span_cols=8)

    # Column headers
    write_label(ws, "B4", "Line Item", bold=True)
    write_units(ws, "C4", "Units")
    for col, label in enumerate(["UW", "Budget", "T-12 Actual", "Var vs. UW", "% Var"], start=4):
        c = ws.cell(row=4, column=col, value=label)
        c.style = "label_bold"
        c.alignment = Alignment(horizontal="right", vertical="center")

    # Rows
    write_label(ws, "B5", "Gross Potential Rent")
    write_units(ws, "C5", "USD")
    write_input(ws, "D5", 5_800_000, fmt="dollar")
    write_input(ws, "E5", 5_850_000, fmt="dollar")
    write_input(ws, "F5", 5_790_000, fmt="dollar")
    write_formula(ws, "G5", "=F5-D5", fmt="dollar")
    write_formula(ws, "H5", "=G5/D5", fmt="pct1")

    write_label(ws, "B6", "Vacancy & Concessions")
    write_units(ws, "C6", "USD")
    write_input(ws, "D6", -290_000, fmt="dollar")
    write_input(ws, "E6", -295_000, fmt="dollar")
    write_input(ws, "F6", -325_000, fmt="dollar")
    write_formula(ws, "G6", "=F6-D6", fmt="dollar")
    write_formula(ws, "H6", "=G6/ABS(D6)", fmt="pct1")

    write_label(ws, "B7", "Other Income")
    write_units(ws, "C7", "USD")
    write_input(ws, "D7", 220_000, fmt="dollar")
    write_input(ws, "E7", 235_000, fmt="dollar")
    write_input(ws, "F7", 245_000, fmt="dollar")
    write_formula(ws, "G7", "=F7-D7", fmt="dollar")
    write_formula(ws, "H7", "=G7/D7", fmt="pct1")

    write_label(ws, "B8", "Effective Gross Income", bold=True)
    for col in range(4, 7):
        write_subtotal(ws, (8, col), f"=SUM({chr(64+col)}5:{chr(64+col)}7)", fmt="dollar")
    write_subtotal(ws, "G8", "=F8-D8", fmt="dollar")
    write_subtotal(ws, "H8", "=G8/D8", fmt="pct1")

    write_label(ws, "B10", "Operating Expenses")
    write_units(ws, "C10", "USD")
    write_input(ws, "D10", -2_300_000, fmt="dollar")
    write_input(ws, "E10", -2_350_000, fmt="dollar")
    write_input(ws, "F10", -2_450_000, fmt="dollar")
    write_formula(ws, "G10", "=F10-D10", fmt="dollar")
    write_formula(ws, "H10", "=G10/ABS(D10)", fmt="pct1")

    write_label(ws, "B12", "Net Operating Income", bold=True)
    for col in range(4, 7):
        write_total(ws, (12, col), f"=SUM({chr(64+col)}8,{chr(64+col)}10)", fmt="dollar")
    write_total(ws, "G12", "=F12-D12", fmt="dollar")
    write_total(ws, "H12", "=G12/D12", fmt="pct1")

    write_section(ws, 14, "Returns Snapshot", span_cols=8)
    write_label(ws, "B15", "Cap Rate (on T-12 NOI)")
    write_units(ws, "C15", "%")
    write_input(ws, "D15", 0.0525, fmt="pct2")

    write_label(ws, "B16", "Implied Value")
    write_units(ws, "C16", "USD")
    write_formula(ws, "D16", "=F12/D15", fmt="dollar")

    write_label(ws, "B17", "MOIC to date")
    write_units(ws, "C17", "x")
    write_input(ws, "D17", 1.42, fmt="multiple")

    write_label(ws, "B18", "Net IRR to date")
    write_units(ws, "C18", "%")
    write_input(ws, "D18", 0.183, fmt="pct1")

    write_note(ws, 20, "Source: T-12 from Yardi as of 2026-03-31; UW from acquisition IC memo dated 2024-08-12.")

    ws.freeze_panes = "D5"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--demo", metavar="OUTPUT", help="Write a styled demo workbook to OUTPUT.xlsx")
    args = ap.parse_args()
    if args.demo:
        _build_demo(args.demo)
        print(f"Wrote demo workbook to {args.demo}")
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
